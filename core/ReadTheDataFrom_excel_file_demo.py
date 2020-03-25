import openpyxl

#  ------------------------------------------------------------------------/ Work book
path_to_excel_file = "C:/Users/User/PycharmProjects/selenium_sdet/test_data/Test.xlsx"
workBook = openpyxl.load_workbook(path_to_excel_file)  # load the workbook (object)
sheet = workBook.active  # if you have only 1 sheet in particular workbook
# workBook.get_sheet_by_name("MyTest_1")
#  ------------------------------------------------------------------------/
rows = sheet.max_row
columns = sheet.max_column
print("Rows is: ", rows)
print("Columns is: ", columns)
#  ------------------------------------------------------------------------/ Reading data from ecxel file
for r in range(1, rows+1):
    for c in range(1, columns+2):
        print(sheet.cell(row=r, column=c).value, end="      ")
    print()
#  ------------------------------------------------------------------------/
#  ------------------------------------------------------------------------/
#  ------------------------------------------------------------------------//
#  ------------------------------------------------------------------------//
#  ------------------------------------------------------------------------//
#  ------------------------------------------------------------------------//
#  ------------------------------------------------------------------------//
#  ------------------------------------------------------------------------//

